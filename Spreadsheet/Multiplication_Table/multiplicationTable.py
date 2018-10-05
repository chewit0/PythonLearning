import openpyxl
import os
import sys, getopt

def create_table(filename, size):
    
    wb = openpyxl.Workbook()
    sheet = wb.active

    for c in range(2,size):
    #     sheet["{}1".format(c)] = c+1
        sheet.cell(row = 1, column = c).value = c-1

    for r in range(2,size):
        sheet["A{}".format(r)] = r-1

    for col in range(2,size):
        for row in range(2,size):
            sheet.cell(row = row, column = col).value =  (row-1) * (col-1)

    wb.save(filename)

def usage():
    print("\nmultiplicationTable.py -o <outputfile> -s <size> \nEnter filename and size of multiplacation table\n")

def main(argv):

    found_filename = False
    found_size = False

    try: 
        opts, args = getopt.getopt(argv, "ho:s:", ["help", "outputfile=", "size="])
    except getopt.GetoptError:
        usage()
        sys.exit()

    for opt, arg in opts:
        if opt in ("-h", "--help"):
            usage()
            sys.exit()
        elif opt in ("-o", "--outputfile"):
            filename = arg
            found_filename = True
        elif opt in ("-s", "--size"):
            try: 
                size = int(arg)
                found_size = True
            except:
                print("You must enter integer for size")
                sys.exit()             

    if found_filename and found_size:
        create_table(filename, size)
    else: 
        usage()

if __name__ == "__main__":
   main(sys.argv[1:])