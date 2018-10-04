import openpyxl
import pprint
import csv
import os

fn = os.path.join(
                  os.path.dirname(__file__), 
                  'Data/'
                  'censuspopdata.xlsx'
                  )

wb = openpyxl.load_workbook(fn)

sheet = wb['Population by Census Tract']
countydata = {}
print(sheet.max_row)

for row in range(2, sheet.max_row):

    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    countydata.setdefault(state, {})
    countydata[state].setdefault(county, {'tracts': 0, 'pop': 0})

    countydata[state][county]['tracts'] += 1
    countydata[state][county]['pop'] += pop

filepath = os.path.join(
                        os.path.dirname(__file__),
                        'condensed_consensus_data.py'
                       )

with open(filepath, 'w') as f:
    f.write('allData = ' + pprint.pformat(countydata))
    f.close()
