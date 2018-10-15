import openpyxl
from openpyxl.styles import Font
import os

'''
Script to read in spreadsheet of produce data and update the cost for specific
items. The updated items are highlighted in bold. A file is created with only
the changed items in.
'''


def rel_fp(path, parent=False, parent_count=0):
    if parent:
        addition = '../'*parent_count
    else:
        addition = ''

    filepath = os.path.join(
                            os.path.dirname(__file__), addition, path
                            )
    return filepath


def produce_update():
    wb = openpyxl.load_workbook(rel_fp('Data/produceSales.xlsx'))

    print(wb.sheetnames)
    data = wb['Sheet']
    max_c = data.max_column

    price_change = {
        'Celery': 1.19,
        'Garlic': 3.07,
        'Lemon': 1.27
    }

    bold = Font(bold=True)

    fpc = rel_fp('changed_produce.xlsx')

    # Create new workbook for changed entries
    try:
        wbc = openpyxl.load_workbook(fpc)
    except:
        wbc = openpyxl.Workbook()
    s_c = wbc['Sheet']

    # Add same column headings
    for x in range(1, max_c+1):
        s_c.cell(row=1, column=x).value = data.cell(1, x).value
    i = 2  # Starting row for new data file, separate index as less rows

    for row in range(2, data.max_row):
        # produce = data['A' + int(row)].value
        produce = data.cell(row, column=1).value

        # Update price if required and bold the row to highlight change.
        if produce in price_change:
            data.cell(row, column=2).value = price_change[produce]        

            for c in range(1, max_c+1):
                data.cell(row, column=c).font = bold

            # Adds changed data to a new file
            for c in range(1, max_c):
                s_c.cell(row=i, column=c).value =
                data.cell(row, column=c).value

            # Adds total column from sum
            s_c.cell(i, max_c).value = '=SUM(B{}*C{})'.format(i, i)

            i += 1  # index for new spreadsheet rows

    filepath_copy = rel_fp('updated_produce.xlsx')

    wb.save(filepath_copy)
    wbc.save(fpc)

if __name__ == "__main__":
    produce_update()
